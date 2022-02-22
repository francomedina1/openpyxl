import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def manipular_xlsx(archivo):
    wb = xl.load_workbook(archivo)
    sheet = wb["Sheet1"]

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        precio_correcto = cell.value * 0.9
        precio_correcto_cell = sheet.cell(row, 4)
        precio_correcto_cell.value = precio_correcto

    values = Reference(sheet, min_row=2, max_row=sheet.max_row,
                    min_col=4, max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, "e2")
    wb.save(archivo)
    
